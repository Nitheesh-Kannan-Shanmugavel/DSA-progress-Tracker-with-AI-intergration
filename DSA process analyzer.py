import sqlite3
import os
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import openai
from pathlib import Path
from dotenv import load_dotenv
import json
import pandas as pd
from matplotlib import pyplot as plt
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

load_dotenv(dotenv_path=Path('data')/'.env')
openai.api_key=os.getenv('OPENAI_API_KEY')
os.makedirs('data',exist_ok=True)

conn=sqlite3.connect('data/dsa_log.db')
cursor=conn.cursor()

# ========== DATABASE SETUP ==========
# All DB initialization code here

def initialize_db():
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS logs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT UNIQUE NOT NULL,
            difficulty TEXT CHECK(difficulty in('Easy','Medium','Hard')),
            platform TEXT,
            date TEXT,
            tags TEXT,
            ai_topic TEXT,
            ai_logic TEXT,
            ai_time_complex TEXT,
            ai_space_complex TEXT,
            company_tags TEXT
        )
            ''')
    conn.commit()
def check_for_duplicates(title:str)->bool:
    cursor.execute("Select 1 from logs where LOWER(title)=lower(?)",(title,))
    return cursor.fetchone() is not None
def get_valid_difficulty():
    while True:
        difficulty = input("Enter difficulty: ").strip().capitalize()
        if difficulty in ('Easy','Medium','Hard'):
            return difficulty
        else:
            print("Invalid difficulty.")
def check_valid_date(date:str)->bool:
    try:
        datetime.strptime(date, '%d/%m/%Y')
        return True
    except ValueError:
        return False

# ========== AI TAGGING LOGIC ==========
# All OpenAI / fallback tagging here / webpage scraping

def scrape_company_tags(title:str)-> list[str] | None:
    try:
        query='+'.join(title.lower().split())
        url=f"https://www.geeksforgeeks.org/?s={query}"
        res=requests.get(url,timeout=10)
        soup=BeautifulSoup(res.text,'html.parser')
        raw_res=soup.select_one('div.head>a')
        if not raw_res:
            raise Exception("No matching GFG problem found.")
        problem_url=raw_res['href']
        problem_request=requests.get(problem_url, timeout=10)
        problem_soup=BeautifulSoup(problem_request.text,'html.parser')
        tags_list=problem_soup.select("a[rel='tag']")
        tags=[tags.text for tags in tags_list if 'Company' in tags.text or 'Company' in tags['href']]
        return tags if tags else None
    except Exception as e:
        print(f'loading failed due to {e}')
        return None
def rule_based_metadata(title):
    title = title.lower()
    if "stock" in title or "buy" in title:
        return {
            "topic": "Greedy",
            "logic": "Track min price and max profit",
            "time": "O(n)",
            "space": "O(1)",
            "company": "Amazon, Google"
        }
    elif "tree" in title:
        return {
            "topic": "Binary Tree",
            "logic": "Use DFS or BFS traversal",
            "time": "O(n)",
            "space": "O(h)",
            "company": "Facebook, Microsoft"
        }
    elif "matrix" in title:
        return {
            "topic": "Matrix Traversal",
            "logic": "Use nested loops or DFS/BFS",
            "time": "O(m*n)",
            "space": "O(1)",
            "company": "Google, Adobe"
        }
    else:
        return {
            "topic": "General DSA",
            "logic": "Based on problem statement",
            "time": "Varies",
            "space": "Varies",
            "company": "Various"
        }


def generate_ai_data(title:str)-> dict:
    company_tags=scrape_company_tags(title)
    if not company_tags:
        prompt=f'''I want you to generate the following list of information for the dsa leetcode problem {title}
            1.DSA logic to solve the problem
            2.time complexity of the problem
            3.space complexity of the problem
            4.companies that have asked this question
            5.topic of the problem
            return strictly in json format like :
            {{
                topic:"...",
                logic : "...",
                time : "...",
                space : "...",
                company: "...",
            }}'''
    else:
        prompt = f'''I want you to generate the following list of information for the dsa leetcode problem {title}
                1.DSA logic to solve the problem
                2.time complexity of the problem
                3.space complexity of the problem
                4.topic of the problem
            return strictly in json format like :
            {{
                topic:"...",
                logic : "...",
                time : "...".
                space : "...".
                
            }}'''
    try:
        response=openai.ChatCompletion.create(
            model='gpt-3.5-turbo',
            messages=[{'role':'user','content':prompt}],
            temperature=0.3,
            max_tokens=200,
        )
        try:
            data=json.loads(response['choices'][0]['message']['content'])
        except json.decoder.JSONDecodeError:
            print("AI doesn't return the file in required format")
            raise Exception('Invalid format')
        normalized={
            'topic':data.get('topic','unknown').strip(),
            'logic':data.get('logic',' ').strip(),
            'time':(data.get('time') or data.get('time complexity','unknown')).strip(),
            'space':(data.get('space') or data.get('space complexity','unknown')).strip(),
            'company': company_tags or data.get('company','unknown')
        }
        return normalized

    except Exception as e:
        print('AI generation failed due to', e)
        print('falling back to rule based ai tagging ')
        fallback=rule_based_metadata(title)
        fallback['company']= ', '.join(company_tags) if company_tags else fallback['company']
        return fallback

# ========== CRUD OPERATIONS ==========
# add_entry, edit_entry, delete_entry, etc.

def add_entry():
    title=input("Enter title: ").strip()
    if not title:
        print("Title cannot be empty")
        return
    if check_for_duplicates(title):
        print("Title already exists")
        return

    difficulty=get_valid_difficulty()
    platform=input("Enter platform: ").strip()
    date=input("Enter date(dd/mm/yyyy): ").strip()
    while not check_valid_date(date):
        print("Invalid date.")
        date=input("Enter date(dd/mm/yyyy): ").strip()
    tags=input("Enter tags(comma separated): ").strip()
    ai_data=generate_ai_data(title)

    cursor.execute("INSERT INTO logs(title,difficulty,platform,date,tags,ai_topic,ai_logic,ai_time_complex,ai_space_complex,company_tags) VALUES (?,?,?,?,?,?,?,?,?,?)",(title,difficulty,platform,date,tags,ai_data['topic'],ai_data['logic'],ai_data['time'],ai_data['space'],ai_data['company']))
    conn.commit()

def view_entry():
    cursor.execute('select * from logs')
    rows=cursor.fetchall()
    if not rows:
        print("No entries found.")
        return
    else:
        for i,entry in enumerate(rows,1):
            print(f"problem {i}")
            print(f"title: {entry[1]}")
            print(f"difficulty: {entry[2]}")
            print(f"platform: {entry[3]}")
            print(f"date: {entry[4]}")
            print(f"tags: {entry[5]}")
            print(f"ai_topic: {entry[6]}")
            print(f"ai_logic: {entry[7]}")
            print(f"ai_time_complex: {entry[8]}")
            print(f"ai_space_complex: {entry[9]}")
            print(f"company_tags: {entry[10]}")

def delete_entry():
    cursor.execute('select * from logs')
    rows=cursor.fetchall()
    if not rows:
        print('No entries found.')
        return
    else:
        for i,entry in enumerate(rows,1):
            print(f"problem {i} - {entry[1]}")
        try:
            choice=int(input("Enter choice: "))
            if 1<=choice<=len(rows):
                cursor.execute('delete from logs where id=?',(rows[choice-1][0],))
                conn.commit()
                print("Deleted entry")
            else:
                print("Invalid choice.Enter the number withing the range")
        except ValueError:
            print("Invalid choice.")
def search_entry():
    cursor.execute('select * from logs')
    rows=cursor.fetchall()
    if not rows:
        print("No entries found.")
        return
    else:
        keyword=input("Enter keyword: ").strip().lower()
        res=[]
        for entry in rows:
            tags_list=[tag.strip().lower() for tag in entry[5].split(',')] if entry[5] else []
            company_list=[company_tags.strip().lower() for company_tags in entry[10].split(',')] if entry[10] else []
            if keyword==entry[1].lower() or keyword==entry[2].lower() or keyword==entry[3].lower() or keyword in tags_list or keyword in company_list or keyword==entry[6].lower() or keyword==entry[7].lower():
                res.append(entry)
        if res:
            for i,entry in enumerate(res,1):
                print(f"problem {i} ")
                print(f"title: {entry[1]}")
                print(f"difficulty: {entry[2]}")
                print(f"platform: {entry[3]}")
                print(f"date: {entry[4]}")
                print(f"tags: {entry[5]}")
                print(f"ai_topic: {entry[6]}")
                print(f"ai_logic: {entry[7]}")
                print(f"ai_time_complex: {entry[8]}")
                print(f"ai_space_complex: {entry[9]}")
                print(f"company_tags: {entry[10]}")
        else:
            print("No entries found.")



def edit_entry():
    cursor.execute('select * from logs')
    rows=cursor.fetchall()
    if not rows:
        print("No entries found.")
        return
    else:
        for i,entry in enumerate(rows,1):
            print(f"problem {i} - {entry[1]}")
        try:
            choice=int(input("Enter your choice of problem to edit :"))
            if 1<=choice<=len(rows):
                oldentry=rows[choice-1]
                newtitle=input("Enter new title: ") or oldentry[1]
                if oldentry[1]!=newtitle and check_for_duplicates(newtitle):
                    print('title already exists')
                    return

                rawdiff=input("Enter new difficulty[Easy,Medium,Hard] : ").capitalize()
                newdiff=rawdiff if rawdiff in ['Easy','Medium','Hard'] else oldentry[2]

                newplatform=input("Enter new platform: ") or oldentry[3]

                rawdate=input("Enter new date: ")
                if rawdate and not check_valid_date(rawdate):
                    print("Invalid date.")
                    return
                newdate=rawdate or oldentry[4]

                rawtags = input("Enter new tags: ").strip()
                newtags = ','.join([tag.strip() for tag in rawtags.split(',')]) if rawtags else oldentry[5]
                cursor.execute('UPDATE logs SET title=?,difficulty=?,platform=?,date=?,tags=? WHERE id=?',(newtitle,newdiff,newplatform,newdate,newtags,oldentry[0]))
                conn.commit()
                print("Edited entry.")
            else:
                print("Invalid entry.")
        except ValueError:
            print("Invalid choice.")

# ========== ANALYTICS ==========
# generate_excel, show_stats, track_targets

def generate_excel(target_count='None',target_date='None'):
    os.makedirs('data',exist_ok=True)
    cursor.execute('select * from logs')
    rows=cursor.fetchall()
    if not rows:
        print("No entries found.")
        return

    df=pd.DataFrame(rows,columns=['ID','title','difficulty','platform','date','tags','ai_topic','ai_logic','ai_time_complex','ai_space_complex','company_tags'])
    df['date']=pd.to_datetime(df['date'],format='%d/%m/%Y')

    data_path='data/DSA_REPORT.xlsx'
    df.to_excel(data_path,index=False)

    topic_count=df['ai_topic'].value_counts()
    plt.figure(figsize=[8,8])
    topic_count.plot.pie(autopct='%1.1f%%',startangle=90,shadow=True)
    plt.title('DSA TOPIC WISE PIE CHART')
    plt.ylabel('')

    buf=BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)

    wb=load_workbook(filename=data_path)
    ws=wb.active
    img=ExcelImage(buf)
    img.anchor='M2'
    ws.add_image(img)

    summary=wb.create_sheet("Summary")
    df['day']=df['date'].dt.date
    df['month']=df['date'].dt.to_period("M")
    day_summary=df.groupby("day").size()
    month_summary=df.groupby("month").size()
    summary.append(['Daily stats'])
    for day,count in day_summary.items():
        summary.append([str(day),count])
    summary.append([])

    summary.append(['Monthly stats'])
    for month,count in month_summary.items():
        summary.append([str(month),count])
    summary.append([])

    today=datetime.today().date()
    target=wb.create_sheet("Target Analysis")
    target.append(['Problems To Reach Target'])
    if target_count:
        remaining=max(0,int(target_count)-len(df))
        print('Problems solved =',len(df))
        print('Problem needs to be solved to reach target = ',remaining)
    if target_date:
        try:
            deadline=datetime.strptime(target_date, "%d/%m/%Y").date()
            days_left=(deadline - today).days
            target.append(['Target Date', target_date])
            target.append(['Days Left', days_left])
        except Exception:
            print("Invalid dateformat")
    wb.save(data_path)
    print('Excel file generated')

# ========== MAIN CLI ==========
# Main loop and menu here

def main():
    initialize_db()
    while True:
        print('Welcome to DSA ANALYSER')
        print('1.add entry')
        print('2.view entries')
        print('3.edit entry')
        print('4.delete entry')
        print('5.search entry')
        print('6.generate excel report')
        print('7.exit')

        choice = input('Enter your choice: ')
        if choice == '1':
            add_entry()
        elif choice == '2':
            view_entry()
        elif choice==  '3':
            edit_entry()
        elif choice == '4':
            delete_entry()
        elif choice == '5':
            search_entry()
        elif choice == '6':
            target_count=input('Enter target count or leave empty if not want to set : ').strip()
            target_date=input('Enter target date or leave empty if not want to set : ').strip()
            target_count=target_count if target_count else None
            target_date=target_date if target_date else None
            generate_excel(target_count,target_date)

        elif choice == '7':
            print('Goodbye')
            break
        else:
            print('Invalid number')

if __name__=="__main__":
    main()
